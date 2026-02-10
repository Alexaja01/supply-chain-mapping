#!/usr/bin/env python3
"""
Excel Import Agent - FIXED for Horizontal Layout
Each row = one terminal with all cost data in columns
"""

import sqlite3
import json
from datetime import datetime, date
import hashlib
import sys
import os

try:
    import openpyxl
except ImportError:
    print("\n❌ ERROR: openpyxl library not installed")
    print("Please install it: pip install openpyxl")
    sys.exit(1)

class ExcelImportAgent:
    """Imports costing data from Costing_Data_Final.xlsx"""
    
    def __init__(self, db_path='supply_chain.db'):
        self.db_path = db_path
        self.effective_date = date(2024, 1, 1)
        
    def import_excel(self, excel_path):
        """Main import workflow"""
        print("📊 Excel Import Agent - Costing Methodology")
        print("=" * 70)
        print(f"  File: {excel_path}")
        print(f"  Effective date: {self.effective_date}")
        print("=" * 70)
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            return {'status': 'failed', 'error': str(e)}
        
        print(f"\n✓ Excel loaded")
        print(f"  Sheets: {workbook.sheetnames}\n")
        
        results = {
            'status': 'completed',
            'terminals_imported': 0,
            'terminal_rates_imported': 0,
            'transport_costs_imported': 0,
        }
        
        # Import from Costing Detail
        if 'Costing Detail' in workbook.sheetnames:
            print("=" * 70)
            print("IMPORTING: Costing Detail")
            print("=" * 70)
            sheet = workbook['Costing Detail']
            
            # Row 3 = headers, data starts row 4
            term_count, rate_count, trans_count = self._import_costing_detail(sheet)
            results['terminals_imported'] = term_count
            results['terminal_rates_imported'] = rate_count
            results['transport_costs_imported'] = trans_count
        
        print("\n" + "=" * 70)
        print("✅ IMPORT COMPLETE!")
        print("=" * 70)
        print(f"  Terminals: {results['terminals_imported']}")
        print(f"  Terminal Rates: {results['terminal_rates_imported']}")
        print(f"  Transportation Costs: {results['transport_costs_imported']}")
        print("=" * 70)
        
        return results
    
    def _import_costing_detail(self, sheet):
        """
        Import from Costing Detail sheet
        Row 3 = Headers
        Row 4+ = Data (each row = one terminal)
        """
        print("\n→ Reading headers from row 3...")
        
        # Get headers from row 3
        headers_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(headers_row)]
        
        print(f"  Found {len(headers)} columns")
        print(f"  Columns A-F: {headers[:6]}")
        
        # Find key columns by index
        state_idx = 0          # Column A
        city_idx = 1           # Column B  
        market_idx = 2         # Column C
        region_idx = 3         # Column D
        code_idx = 4           # Column E
        name_idx = 5           # Column F
        
        # Find product columns (Clear Gas, E10, E15)
        product_columns = {}
        for i, header in enumerate(headers):
            if header == 'Clear Gas':
                product_columns['Clear Gas'] = i
            elif header == 'E10':
                product_columns['E10'] = i
            elif header == 'E15':
                product_columns['E15'] = i
        
        print(f"  Product columns found: {product_columns}")
        
        print("\n→ Processing terminal rows (starting row 4)...")
        
        terminals = []
        terminal_rates = []
        transport_costs = []
        
        row_num = 0
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not row or not any(row):
                continue
            
            row_num += 1
            
            # Extract terminal info
            state = row[state_idx] if state_idx < len(row) else None
            city = row[city_idx] if city_idx < len(row) else None
            market = row[market_idx] if market_idx < len(row) else None
            region = row[region_idx] if region_idx < len(row) else None
            terminal_code = row[code_idx] if code_idx < len(row) else None
            terminal_name = row[name_idx] if name_idx < len(row) else None
            
            # Skip if missing key fields
            if not state or not city or not terminal_name:
                continue
            
            # Generate terminal ID
            terminal_id = self._generate_terminal_id(state, city, terminal_code)
            
            # Store terminal
            terminals.append({
                'terminal_id': terminal_id,
                'terminal_name': terminal_name,
                'terminal_code': terminal_code,
                'state': state,
                'city': city,
                'market': market,
                'region': region
            })
            
            # For each product, extract costs
            for product_name, col_idx in product_columns.items():
                if col_idx < len(row):
                    # Get the combined adder value for this product
                    combined_adder = self._parse_number(row[col_idx])
                    
                    if combined_adder is not None:
                        # Store transportation cost with combined adder
                        trans_id = self._generate_transport_id(terminal_id, product_name)
                        transport_costs.append({
                            'transport_cost_id': trans_id,
                            'terminal_id': terminal_id,
                            'product_type': product_name,
                            'combined_adder': combined_adder
                        })
            
            if row_num % 50 == 0:
                print(f"  Processed {row_num} rows...")
        
        print(f"\n  Total rows processed: {row_num}")
        print(f"  Unique terminals: {len(terminals)}")
        print(f"  Transport cost records: {len(transport_costs)}")
        
        # Store in database
        print("\n→ Storing in database...")
        term_count = self._store_terminals(terminals)
        trans_count = self._store_transport_costs(transport_costs)
        
        return term_count, 0, trans_count
    
    def _store_terminals(self, terminals):
        """Store terminals in database"""
        print("  → Storing terminals...")
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        count = 0
        for terminal in terminals:
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO terminals (
                        terminal_id, terminal_name, terminal_code, state, city,
                        market, region, effective_date, data_quality_score,
                        created_by, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    terminal['terminal_id'],
                    terminal['terminal_name'],
                    terminal.get('terminal_code'),
                    terminal['state'],
                    terminal['city'],
                    terminal.get('market'),
                    terminal.get('region'),
                    self.effective_date,
                    0.95,
                    'excel_import_agent',
                    datetime.now(),
                    datetime.now()
                ))
                count += 1
            except Exception as e:
                print(f"    ⚠️  Error: {terminal.get('terminal_name')}: {e}")
        
        conn.commit()
        conn.close()
        
        print(f"    ✓ Stored {count} terminals")
        return count
    
    def _store_transport_costs(self, transport_costs):
        """Store transportation costs"""
        print("  → Storing transportation costs...")
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        count = 0
        for cost in transport_costs:
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO transportation_costs (
                        transport_cost_id, terminal_id, product_type,
                        combined_adder, effective_date,
                        created_by, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    cost['transport_cost_id'],
                    cost['terminal_id'],
                    cost['product_type'],
                    cost['combined_adder'],
                    self.effective_date,
                    'excel_import_agent',
                    datetime.now(),
                    datetime.now()
                ))
                count += 1
            except Exception as e:
                print(f"    ⚠️  Error storing cost: {e}")
        
        conn.commit()
        conn.close()
        
        print(f"    ✓ Stored {count} transport costs")
        return count
    
    def _generate_terminal_id(self, state, city, terminal_code):
        """Generate unique terminal ID"""
        if terminal_code:
            return f"TERM_{state}_{terminal_code}"
        else:
            combined = f"{state}_{city}".lower().replace(" ", "_")
            hash_val = hashlib.md5(combined.encode()).hexdigest()[:8]
            return f"TERM_{state}_{hash_val}"
    
    def _generate_transport_id(self, terminal_id, product):
        """Generate unique transport cost ID"""
        combined = f"{terminal_id}_{product}".lower().replace(" ", "_")
        hash_val = hashlib.md5(combined.encode()).hexdigest()[:8]
        return f"TRANS_{hash_val}"
    
    def _parse_number(self, value):
        """Parse number from various formats"""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace('$', '').replace(',', '').replace(' ', '').strip()
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

# Command-line interface
if __name__ == "__main__":
    print("\n" + "="*80)
    print("  EXCEL IMPORT AGENT - COSTING METHODOLOGY")
    print("  Imports 227 terminals from Costing_Data_Final.xlsx")
    print("="*80 + "\n")
    
    # Try to find Excel file
    possible_paths = [
        os.path.join('Reference', 'Excel', 'Costing_Data_Final.xlsx'),
        os.path.join('reference', 'excel', 'Costing_Data_Final.xlsx'),
    ]
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = None
        print("Searching for Excel file...")
        for path in possible_paths:
            print(f"  Checking: {path} ... ", end="")
            if os.path.exists(path):
                excel_path = path
                print("✓")
                break
            else:
                print("✗")
        
        if not excel_path:
            excel_path = possible_paths[0]
    
    if not os.path.exists(excel_path):
        print(f"\n❌ Excel file not found: {excel_path}")
        print("\nUsage:")
        print("  python excel_import_agent.py [path/to/file.xlsx]")
        sys.exit(1)
    
    # Run import
    agent = ExcelImportAgent('supply_chain.db')
    results = agent.import_excel(excel_path)
    
    if results['status'] == 'completed' and results['terminals_imported'] > 0:
        print("\n✅ SUCCESS!")
        print(f"\nImported {results['terminals_imported']} terminals!")
        print("\nNext steps:")
        print("  python orchestrator.py --api-key YOUR_KEY status")
    else:
        print(f"\n❌ Import completed but no data imported")
        print("Check if Excel structure matches expected format")
        sys.exit(1)
